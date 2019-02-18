from openpyxl import workbook, worksheet
from requests import get
from json import loads
from time import strftime
from shutil import copyfile, move

COLUMN_WIDTH = 50

config = loads(open("config.json").read())

adminAuth = config["database secret"]
basePath = config["database base url"]


fileName = strftime("%Y%m%d-%H")
try:
    move("numbers_sync.xlsx", "backups_sync/{}.xlsx".format(fileName))
except:
    print('error in backup', 'no file or permission denied')

wb = workbook.Workbook()
wb.remove(wb.active)

sheetNames = ('we vip', 'we مميز', 'اتصالات vip', 'اتصالات مميز', 'فودافون vip', 'فودافون مميز', 'اورانج vip', 'اورانج مميز')
dbLocations = ("We-vip", "We-special", "Etisalat-vip", "Etisalat-special", "Vodafone-vip", "Vodafone-special", "Orange-vip", "Orange-special")

for sheetName, dbLocation in zip(sheetNames, dbLocations):
    url = "{}/{}.json?auth={}".format(basePath, dbLocation, adminAuth)
    numbers = get(url).json()
    if(not numbers):
        continue
    sheet = wb.create_sheet(title=sheetName)
    sheet.column_dimensions["A"].width = COLUMN_WIDTH
    iRow = 1
    for number in numbers:
        sheet.cell(row=iRow, column=1, value=number).style = "Normal" if numbers[number] == "available" else "Bad"
        iRow += 1

dst = "./numbers_sync.xlsx"
wb.save(dst)
wb.close()