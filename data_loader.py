from openpyxl import load_workbook
from requests import put
from json import dumps, loads
from time import strftime
from shutil import copyfile

config = loads(open("config.json").read())

adminAuth = config["database secret"]
basePath = config["database base url"]

src = "./numbers.xlsx"
wb = load_workbook(src)

sheetNames = ('we vip', 'we مميز', 'اتصالات vip', 'اتصالات مميز', 'فودافون vip', 'فودافون مميز', 'اورانج vip', 'اورانج مميز')
dbLocations = ("We-vip", "We-special", "Etisalat-vip", "Etisalat-special", "Vodafone-vip", "Vodafone-special", "Orange-vip", "Orange-special")


for sheetName, dbLocation in zip(sheetNames, dbLocations):
    sheet = wb[sheetName]
    count = sheet.max_row
    print(sheet.title, count)
    numbers = {}
    for i in range(count):
        currCell = sheet[i + 1][0]
        phoneNumber = currCell.value
        status = "unavailable" if currCell.style == "Bad" else "available"
        if (phoneNumber == None or phoneNumber.strip() == ""):
            continue
        numbers[phoneNumber] = status
    url = "{}/{}.json?auth={}".format(basePath, dbLocation, adminAuth)
    res = put(url, data=dumps(numbers))
    print("res for", url, "is", res)

wb.close()

# fileName = strftime("%Y%m%d-%H%M%S")
fileName = strftime("%Y%m%d-%H")
copyfile(src, "backups/{}.xlsx".format(fileName))